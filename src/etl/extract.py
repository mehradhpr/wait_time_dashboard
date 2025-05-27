
# Data extraction utilities for various sources

import pandas as pd
import logging
from pathlib import Path
from typing import Dict, List, Optional, Union
import openpyxl
from datetime import datetime

logger = logging.getLogger(__name__)

class DataExtractor:
    """Base class for data extraction operations"""
    
    def __init__(self):
        self.extraction_stats = {
            'source_file': None,
            'records_extracted': 0,
            'extraction_time': None,
            'errors': []
        }
    
    def log_extraction_stats(self):
        """Log extraction statistics"""
        logger.info(f"Extraction completed for {self.extraction_stats['source_file']}")
        logger.info(f"Records extracted: {self.extraction_stats['records_extracted']}")
        logger.info(f"Extraction time: {self.extraction_stats['extraction_time']}")
        if self.extraction_stats['errors']:
            logger.warning(f"Extraction errors: {len(self.extraction_stats['errors'])}")

class ExcelExtractor(DataExtractor):
    """Excel file extraction with healthcare wait times specific logic"""
    
    def extract_wait_times_data(self, file_path: Union[str, Path]) -> pd.DataFrame:
        """Extract wait times data from the specific Excel format"""
        start_time = datetime.now()
        self.extraction_stats['source_file'] = str(file_path)
        
        try:
            logger.info(f"Extracting data from {file_path}")
            
            # Validate file exists
            if not Path(file_path).exists():
                raise FileNotFoundError(f"Source file not found: {file_path}")
            
            # Read the specific worksheet with wait time data
            df = pd.read_excel(
                file_path,
                sheet_name='Wait times 2008 to 2023',
                skiprows=2,  # Skip the header information rows
                engine='openpyxl',
                na_values=['n/a', 'N/A', '', ' ']
            )
            
            # Basic data validation
            if df.empty:
                raise ValueError("Extracted dataframe is empty")
            
            # Clean column names
            df.columns = df.columns.str.strip()
            
            # Remove completely empty rows
            df = df.dropna(how='all')
            
            # Filter out rows where key columns are all empty
            key_columns = ['Province/territory', 'Indicator', 'Metric', 'Data year']
            df = df.dropna(subset=key_columns, how='all')
            
            # Remove extra empty columns (common in Excel exports)
            df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
            
            # Record extraction stats
            self.extraction_stats['records_extracted'] = len(df)
            self.extraction_stats['extraction_time'] = datetime.now() - start_time
            
            logger.info(f"Successfully extracted {len(df)} records")
            self.log_extraction_stats()
            
            return df
            
        except Exception as e:
            error_msg = f"Data extraction failed: {str(e)}"
            self.extraction_stats['errors'].append(error_msg)
            logger.error(error_msg)
            raise
    
    def extract_sheet_metadata(self, file_path: Union[str, Path]) -> Dict:
        """Extract metadata about the Excel file structure"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            metadata = {
                'file_path': str(file_path),
                'sheet_names': workbook.sheetnames,
                'sheet_info': {}
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                metadata['sheet_info'][sheet_name] = {
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                    'title': sheet.title
                }
            
            workbook.close()
            return metadata
            
        except Exception as e:
            logger.error(f"Failed to extract sheet metadata: {e}")
            return {}

class CSVExtractor(DataExtractor):
    """CSV file extraction utilities"""
    
    def extract_csv_data(self, file_path: Union[str, Path], 
                        encoding: str = 'utf-8',
                        delimiter: str = ',') -> pd.DataFrame:
        """Extract data from CSV files"""
        start_time = datetime.now()
        self.extraction_stats['source_file'] = str(file_path)
        
        try:
            logger.info(f"Extracting CSV data from {file_path}")
            
            df = pd.read_csv(
                file_path,
                encoding=encoding,
                delimiter=delimiter,
                low_memory=False
            )
            
            self.extraction_stats['records_extracted'] = len(df)
            self.extraction_stats['extraction_time'] = datetime.now() - start_time
            
            logger.info(f"Successfully extracted {len(df)} records from CSV")
            return df
            
        except Exception as e:
            error_msg = f"CSV extraction failed: {str(e)}"
            self.extraction_stats['errors'].append(error_msg)
            logger.error(error_msg)
            raise

class DatabaseExtractor(DataExtractor):
    """Database extraction utilities"""
    
    def __init__(self, db_connection):
        super().__init__()
        self.db_connection = db_connection
    
    def extract_table_data(self, table_name: str, 
                          columns: Optional[List[str]] = None,
                          where_clause: Optional[str] = None) -> pd.DataFrame:
        """Extract data from database table"""
        start_time = datetime.now()
        self.extraction_stats['source_file'] = f"database_table_{table_name}"
        
        try:
            # Build query
            column_list = ', '.join(columns) if columns else '*'
            query = f"SELECT {column_list} FROM {table_name}"
            
            if where_clause:
                query += f" WHERE {where_clause}"
            
            logger.info(f"Extracting data from table: {table_name}")
            
            # Execute query
            result = self.db_connection.execute_query(query, fetch=True)
            
            # Convert to DataFrame
            df = pd.DataFrame(result)
            
            self.extraction_stats['records_extracted'] = len(df)
            self.extraction_stats['extraction_time'] = datetime.now() - start_time
            
            logger.info(f"Successfully extracted {len(df)} records from {table_name}")
            return df
            
        except Exception as e:
            error_msg = f"Database extraction failed: {str(e)}"
            self.extraction_stats['errors'].append(error_msg)
            logger.error(error_msg)
            raise

def extract_wait_times_data(file_path: Union[str, Path]) -> pd.DataFrame:
    """Convenience function for extracting wait times data"""
    extractor = ExcelExtractor()
    return extractor.extract_wait_times_data(file_path)

def validate_data_structure(df: pd.DataFrame) -> Dict[str, bool]:
    """Validate that extracted data has expected structure"""
    expected_columns = [
        'Reporting level', 'Province/territory', 'Region',
        'Indicator', 'Metric', 'Data year', 
        'Unit of measurement', 'Indicator result'
    ]
    
    validation_results = {
        'has_required_columns': all(col in df.columns for col in expected_columns),
        'has_data': len(df) > 0,
        'has_valid_years': False,
        'has_valid_provinces': False
    }
    
    if validation_results['has_data']:
        # Check for valid years
        year_column = 'Data year'
        if year_column in df.columns:
            years = pd.to_numeric(df[year_column], errors='coerce')
            validation_results['has_valid_years'] = years.between(2008, 2023).any()
        
        # Check for valid provinces
        province_column = 'Province/territory'
        if province_column in df.columns:
            provinces = df[province_column].dropna().unique()
            canadian_provinces = ['Alberta', 'British Columbia', 'Manitoba', 'Ontario', 'Quebec']
            validation_results['has_valid_provinces'] = any(prov in provinces for prov in canadian_provinces)
    
    return validation_results